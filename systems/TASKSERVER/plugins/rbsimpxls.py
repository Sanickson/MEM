# -*- coding: utf-8 -*-
"""
    модуль импорта данных XLS
    базовый модуль
"""
import openpyxl
import os
import re
import BasePlugin as BP
import krconst
# import datetime


class Plugin(BP.BasePlugin):
    """
        модуль импорта данных XLS
        базовый модуль
    """

    def dataPrep(self, data):
        if type(data) is str:
            data = data.replace(' ', '').strip()
        if data == '-':
            data = None
        return None if data == '' else data

    def floatPrep(self, data):
        data = self.dataPrep(data)
        if data:
            data = str(data).replace(',','.').replace(' ','')
        return data

    def run(self):
        # проверим формат файла
        ReferenceGwares = re.findall(r'xls', os.path.basename(self.filenames))
        if(len(ReferenceGwares) > 0):
            # открываем документ
            book = openpyxl.load_workbook(filename=self.filenames, read_only=True)
            # берем первую страницу
            sheet=book.worksheets[0]
            # проверка на содержимое файла
            if (sheet.cell(2,2).value == "КОД" and sheet.cell(2,3).value == "Название" and sheet.cell(2,4).value == "Адрес"):
                self.departReference(sheet)
            elif (re.findall(r'Постачальники по відділах', str(sheet.cell(1,1).value))):
                self.objReference(sheet)
            elif (re.findall(r'Оборотно-сальдовая ведомость по счету', str(sheet.cell(2,2).value))):
                self.objReference2(sheet)
            elif (re.findall(r'Приходная накладная', str(sheet.cell(4,1).value))):
                self.incomingDocument(sheet)
            elif (re.findall(r'Оприходования остатков', str(sheet.cell(3,1).value))):
                self.incomingDocument2(sheet)
            elif (re.findall(r'Партионное межмаркетовое перемещение товара', str(sheet.cell(1, 5).value))):
                self.saleDocument(sheet)
            elif (sheet.cell(1,1).value == "Артикул" and sheet.cell(1,2).value == "Наименование" and sheet.cell(1,3).value == "Отдел"):
                self.waresReference(sheet)    
            elif (sheet.cell(1,1).value == "Розширенний список товарів"):
                self.extraWaresReference(sheet)
            else:
                raise Exception("Файл не обрабатывается")
        """
        # создадим переменные
        self.code = ""
        self.name = ""
        self.waresgroup = ""
        # проверим формат файла
        if (os.path.basename(self.filenames)).split('.')[1] in ('xls', 'xlsx'):
            # проверка файла по названию
            ReferenceGwares = re.findall(r'Справочник товаров', self.filenames)
            if (len(ReferenceGwares) > 0):
                # открываем документ
                book = xlrd.open_workbook(filename=self.filenames, on_demand=True)
                # для всех страниц в документе
                for i in range(book.nsheets):
                    sheet = book.sheet_by_index(i)
                    # для всех строк на странице
                    for num_row in range(sheet.nrows):
                        # для всех столбцов на странице
                        for num_col in range(sheet.ncols):
                            # находим оглавние для каждого столбца
                            cellMain = sheet.cell(0, num_col)
                            if (cellMain.value == 'Артикул'):
                                cell = sheet.cell(num_row+1, num_col)
                                self.code = cell.value
                            if (cellMain.value == 'Наименование'):
                                cell = sheet.cell(num_row+1, num_col)
                                self.name = cell.value
                            if (cellMain.value == 'Группа'):
                                cell = sheet.cell(num_row+1, num_col)
                                self.waresgroup = cell.value
                        try:
                            sql_text = 'execute procedure TS_IMPORTXLS_GWARES(?,?,?)'
                            sql_params = [self.name, self.code, self.waresgroup]
                            self.db.dbExec(sql_text,
                                           params=sql_params,
                                           fetch='None')
                        except Exception:
                            return "Ошибка добавления данных в БД"
                            """

    def correctPhone(self, phone, num):
        correctPhone = re.split(r',|;|/', phone)[num]
        correctPhone = ''.join(re.findall(r'\d', correctPhone))
        correctPhone = re.findall(r'\d{7,12}$', correctPhone)
        if (correctPhone == []):
            correctPhone = None
        elif (correctPhone != []):
            correctPhone = correctPhone[0]

        return correctPhone

    def departReference(self, sheet):
        ir = sheet.iter_rows(min_row=3, max_row=sheet.max_row, min_col=2, max_col=sheet.max_column, values_only=True)
        for item in ir:
            code = item[0]
            name = item[1]
            address = item[2]

            sql_text = 'execute procedure TS_IMPXLS_DEPART(?,?,?)'
            sql_params = [name, code, address]
            self.ExecuteSQL(sql_text,
                            sqlparams=sql_params,
                            fetch='none',
                            ExtVer=True)
                            
                            
    def updateWaresReference(self, sheet):
        ir = sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column, values_only=True)
        for item in ir:
            code = item[0]
            name = item[1]
            waresgrouproot = item[4]
            waresgroup = item[5]
            selgroup = item[2]
            sql_text = 'execute procedure TS_IMPXLS_CHANGE_SELGROUP(?,?,?,?,?)'
            sql_params = [code, name, waresgrouproot, waresgroup, selgroup]
            self.ExecuteSQL(sql_text,
                            sqlparams=sql_params,
                            fetch='none',
                            ExtVer=True)

    def waresReference(self, sheet):
        ir = sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column, values_only=True)
        for item in ir:
            code = item[0]
            name = item[1]
            waresgrouproot = item[4]
            waresgroup = item[5]
            sql_text = 'execute procedure TS_IMPXLS_GWARES(?,?,?,?)'
            sql_params = [code, name, waresgrouproot, waresgroup]
            self.ExecuteSQL(sql_text,
                            sqlparams=sql_params,
                            fetch='none',
                            ExtVer=True)

    def extraWaresReference(self, sheet):
        ir = sheet.iter_rows(min_row=3, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column, values_only=True)
        for item in ir:
            code = self.dataPrep(item[2])
            barcode = self.dataPrep(item[3])
            name = self.dataPrep(item[4])
            if code or barcode or name:
                factor = self.dataPrep(item[5])
                if factor:
                    factor = factor.replace(',','.').replace(' ','')
                unit = self.dataPrep(item[8]) # уп. = упак, м.п = метр
                supplier = None #self.dataPrep(item[9])
                producer = None #self.dataPrep(item[13])

                sql_text = 'execute procedure TS_IMPXLS_EXTRAGWARES(?,?,?,?,?,?,?)'
                sql_params = [code,name,producer, supplier, unit, factor, barcode]
                self.ExecuteSQL(sql_text,
                                sqlparams=sql_params,
                                fetch='none',
                                ExtVer=True)


    def objReference(self, sheet):
        ir = sheet.iter_rows(min_row=3, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column, values_only=True)
        for item in ir:
            name = self.dataPrep(item[2])
            if name:
                address = self.dataPrep(item[3])
                okpo = self.dataPrep(item[4])
                phone = self.dataPrep(item[5])
                if phone is None:
                    mainphone = None
                    extraphone = None
                elif (re.findall(r',|;|/', phone)):
                    mainphone = self.correctPhone(phone, 0)
                    extraphone = self.correctPhone(phone, 1)
                elif (''.join(re.findall(r'\d', phone)) != ''):
                    mainphone = ''.join(re.findall(r'\d', phone))
                    extraphone = None
                else:
                    mainphone = phone
                    extraphone = None
                sql_text = 'execute procedure TS_IMPXLS_OBJ(?,?,?,?,?)'
                sql_params = [name, okpo, address, mainphone, extraphone]
                self.ExecuteSQL(sql_text,
                                sqlparams=sql_params,
                                fetch='none',
                                ExtVer=True)


    def objReference2(self, sheet):
        name_item = 9
        okpo_item = 10
        while (name_item <= sheet.max_row or okpo_item <= sheet.max_row):
            name = self.dataPrep(sheet.cell(row=name_item, column=2).value)
            okpo = self.dataPrep(sheet.cell(row=okpo_item, column=2).value)
            sql_text = 'execute procedure TS_IMPXLS_OBJ(?,?,?,?,?)'
            sql_params = [name, okpo, None, None, None]
            self.ExecuteSQL(sql_text,
                            sqlparams=sql_params,
                            fetch='none',
                            ExtVer=True)
            name_item += 2
            okpo_item += 2

    def incomingDocument(self, sheet):
        fromobj = sheet.cell(1,1).value.split("\n")[0].split(": ")[1]
        toobj = sheet.cell(5,1).value.split(": ")[1]
        doctype = 'INCOME'
        nd = sheet.cell(4,1).value.split('от')
        docdate = nd[1].strip()
        docnumber = nd[0].split('№')[1].strip()

        sql_text = 'select * from TS_IMPXLS_ADDDOC(?,?,?,?,?)'
        sql_params = [docnumber, docdate, fromobj, toobj, doctype]
        docres = self.ExecuteSQL(sql_text,
                        sqlparams=sql_params,
                        fetch='one',
                        ExtVer=True)
        if docres['status'] == krconst.kr_sql_error:
            return
        docid = docres['datalist']['DOCID']

        flagEnd = True
        ir = sheet.iter_rows(min_row=14, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column, values_only=True)
        for item in ir:
            num = self.dataPrep(item[0])
            code = self.dataPrep(item[1])
            name = self.dataPrep(item[3])
            if num and code and name:
                barcode = self.dataPrep(item[2])
                unit = self.dataPrep(item[5])
                amount = self.floatPrep(item[6])
                summa = self.floatPrep(item[9])
                cgres = self.ExecuteSQL('execute procedure TS_IMPXLS_DOCWARES(?,?,?,?,?,?,?)',
                                        sqlparams=[docid, code, name, barcode, unit, amount, summa],
                                        fetch='none',
                                        ExtVer=True)
                if cgres['status'] == krconst.kr_sql_error:
                    flagEnd = False

        if flagEnd:
            self.ExecuteSQL('execute procedure TS_IMPXLS_DOCEND(?)', sqlparams=[docid], fetch='one', ExtVer=True)


    def incomingDocument2(self, sheet):
        fromobj = sheet.cell(1,1).value.strip()
        toobj = sheet.cell(2,1).value.strip()
        doctype = 'INCOME'
        nd = sheet.cell(3,1).value.split('от')
        docdate = nd[1].strip()
        docnumber = nd[0].split('№')[1].strip()

        sql_text = 'select * from TS_IMPXLS_ADDDOC(?,?,?,?,?)'
        sql_params = [docnumber, docdate, fromobj, toobj, doctype]
        docres = self.ExecuteSQL(sql_text,
                        sqlparams=sql_params,
                        fetch='one',
                        ExtVer=True)
        if docres['status'] == krconst.kr_sql_error:
            return
        docid = docres['datalist']['DOCID']

        flagEnd = True
        ir = sheet.iter_rows(min_row=5, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column, values_only=True)
        for item in ir:
            num = self.dataPrep(item[0])
            code = self.dataPrep(item[1])
            name = self.dataPrep(item[2])
            amount = self.floatPrep(item[5])
            if num and code and name and float(amount) > 0:
                unit = self.dataPrep(item[3])
                summa = self.floatPrep(item[7])
                cgres = self.ExecuteSQL('execute procedure TS_IMPXLS_DOCWARES(?,?,?,?,?,?,?)',
                                        sqlparams=[docid, code, name, None, unit, amount, summa],
                                        fetch='none',
                                        ExtVer=True)
                if cgres['status'] == krconst.kr_sql_error:
                    flagEnd = False

        if flagEnd:
            self.ExecuteSQL('execute procedure TS_IMPXLS_DOCEND(?)', sqlparams=[docid], fetch='one', ExtVer=True)


    def objName(self, cellvalue, splitsymbol=':'):
        spl = cellvalue.split(':')
        if len(spl)>1:
            return splitsymbol.join(spl[1:]).strip()
        else:
            return cellvalue.strip()

    def saleDocument(self, sheet):
        fromobj = self.objName(sheet.cell(1,1).value.split('_x000a_')[4])
        toobj = self.objName(cellvalue=sheet.cell(1,6).value.split('_x000a_')[0])
        doctype = 'SALE'
        nd = sheet.cell(1,5).value.split('от')
        docdate = nd[1].strip().split(' ')[0]
        docnumber = nd[0].split('№')[1].strip()

        sql_text = 'select * from TS_IMPXLS_ADDDOC(?,?,?,?,?)'
        sql_params = [docnumber, docdate, fromobj, toobj, doctype]
        docres = self.ExecuteSQL(sql_text,
                        sqlparams=sql_params,
                        fetch='one',
                        ExtVer=True)
        if docres['status'] == krconst.kr_sql_error:
            return
        docid = docres['datalist']['DOCID']

        flagEnd = True
        ir = sheet.iter_rows(min_row=3, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column, values_only=True)
        i = 1
        for item in ir:
            num = self.dataPrep(item[0])
            if num == 'Контроль затронутых документов-перемещений':
                break
            code = self.dataPrep(item[1])
            name = self.dataPrep(item[4])
            if num and num == i and code and name:
                barcode = self.dataPrep(item[2])
                unit = self.dataPrep(item[5])
                amount = self.floatPrep(item[6])
                summa = self.floatPrep(item[9])
                cgres = self.ExecuteSQL('execute procedure TS_IMPXLS_DOCWARES(?,?,?,?,?,?,?)',
                                        sqlparams=[docid, code, name, barcode, unit, amount, summa],
                                        fetch='none',
                                        ExtVer=True)
                if cgres['status'] == krconst.kr_sql_error:
                    flagEnd = False
                i += 1

        if flagEnd:
            self.ExecuteSQL('execute procedure TS_IMPXLS_DOCEND(?)', sqlparams=[docid], fetch='one', ExtVer=True)
